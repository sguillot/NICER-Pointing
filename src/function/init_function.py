# ------------------------------ #
        # Python's packages
        
from astropy.io import fits
from astropy.table import Table
from termcolor import colored
from astropy.coordinates import SkyCoord
from typing import Dict, Tuple, List
from astroquery.simbad import Simbad
from astropy import units as u

import os
import argparse
import numpy as np
import openpyxl

# ------------------------------ #

"""
This module provides a collection of functions primarily aimed at handling astronomical data and workflows. It integrates functionalities from several packages such as Astropy, Astroquery, NumPy, Openpyxl, and others, to facilitate various operations like validating file paths, interacting with astronomical catalogs, and managing data in Python and Excel formats.

Functions:

- is_valid_file_path(file_path): Checks if a file exists at a given file path.
- get_valid_file_path(path): Continuously prompts the user for a valid file path until one is provided.
- choose_catalog(args_catalog): Chooses a catalog based on a provided keyword and returns a valid file path for it.
- define_sources_list(): Prompts the user to define a list of sources for calculations, either manually or by importing from a file.
- add_source_list(active_workflow): Prompts the user to add sources to a calculation and loads a FITS file as a source list.
- get_coord_psr(name): Retrieves the coordinates of a pulsar object from the SIMBAD database.
- py_to_xlsx(excel_data_path, count_rates, object_data, args, radius): Converts and saves Python data into an Excel file.
- xlsx_to_py(excel_data_path, nearby_sources_table, object_data, args, radius): Reads count rate data from an Excel file and integrates it into a Python table.

The module's functionality is diverse, catering to specific needs in astronomical data analysis and handling. It emphasizes user interaction for data input and validation, ensuring that the operations are carried out on correct and existing data paths. The integration with astronomical catalogs and databases like SIMBAD, as well as the functionality to convert between Python data structures and Excel files, makes it a versatile tool for astronomers and data analysts working in related fields.
"""



def is_valid_file_path(file_path) -> bool:
    """
    Determines whether a file exists at the specified file path.

    This function checks the existence of a file at a given path. It's useful for validating file paths 
    before attempting to open or process files, thereby preventing errors related to file access or non-existence.

    Args:
        file_path (str): The path of the file to be checked.

    Returns:
        bool: Returns True if the file exists at the given path, False otherwise.

    The function uses the os.path.exists method to verify the existence of the file. In case of any 
    unexpected error during the check, the error is caught and printed, and the function returns False.

    Note:
        - The function is designed to handle general file path checks and is not specific to any file type.
        - If the file path is invalid or inaccessible due to permissions, the function returns False.
    """
    try:
        if os.path.exists(file_path):
            return True
        else:
            return False
    except Exception as error:
        print(f"An error occured: {error}")


def get_valid_file_path(path) -> str:
    """
    Prompt the user for a valid file path until a valid one is provided.

    Args:
        path (str): The initial file path.

    Returns:
        str: A valid file path that exists.
    """
    while True :
        if is_valid_file_path(path):
            print(f"The file at {colored(path, 'yellow')} is {colored('valid', 'green')}.")
            break
        else:
            print(f"The file at {colored(path, 'yellow')} doesn't exist or the path is {colored('invalid', 'red')}.")
            path = str(input("Enter the file path : \n"))
    return path


def choose_catalog(args_catalog) -> Tuple[str, str]:
    """
    Choose a catalog based on the provided keyword and return a valid file path for it.

    Args:
        Catalog (str): The catalog keyword, should be 'DR11' or 'DR13'.

    Returns:
        str: A valid file path for the selected catalog.

    Raises:
        argparse.ArgumentError: If an invalid catalog keyword is provided.
    """
    active_workflow = os.getcwd().replace("\\", "/")
    catalog_datapath = os.path.join(active_workflow, "data/catalog_data").replace("\\", "/")
    while True:
        try:
            if args_catalog == 'Xmm_DR13':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                xmm_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(xmm_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'CSC_2.0':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                chandra_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(chandra_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'Swift':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                swift_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(swift_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'eRosita':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                eRosita_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(eRosita_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == "compare_catalog":
                print("\n")
                print("Enter catalog keyword (Xmm_DR13/CSC_2.0/Swift/eRosita)")
                catalog_1 = str(input("First catalog : "))
                while True:
                    if catalog_1 == "Xmm_DR13":
                        catalog_1_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "CSC_2.0":
                        catalog_1_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "Swift":
                        catalog_1_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "eRosita":
                        catalog_1_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                        break
                    else:
                        catalog_1 = str(input("Keyword unnaccepted, retry : "))
                valid_path_1 = get_valid_file_path(catalog_1_path)
                
                catalog_2 = str(input("Second catalog : "))
                while True:
                    if catalog_2 == "Xmm_DR13":
                        catalog_2_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "CSC_2.0":
                        catalog_2_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "Swift":
                        catalog_2_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "eRosita":
                        catalog_2_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                        break
                    else:
                        catalog_2 = str(input("Keyword unnaccepted, retry : "))
                valid_path_2 = get_valid_file_path(catalog_2_path)
                
                valid_path = (valid_path_1, valid_path_2, catalog_1, catalog_2)
                return valid_path, args_catalog
            elif args_catalog == "match":
                return 'matched_catalog', 'match'
            else:
                raise argparse.ArgumentError(None, "invalid catalog keyword keyword. retry with Xmm_DR13, CSC_2.0, Swift, eRosita, compare_catalog or match.")
        except argparse.ArgumentError as error:
            print(f'An error occured : {error}')
            args_catalog = str(input("Enter a new key word : \n"))


def define_sources_list() -> Table:
    """
    Prompts the user to define a list of astronomical sources for calculations. This function allows users to add sources either manually or by importing data from a file.

    The function initially asks the user if they wish to add sources to the calculation. Users can add sources manually by entering the source details (name, right ascension (ra), declination (dec), and variability rate) or choose to import this data from a specified file. The data, whether manually input or imported, is stored as a list of tuples, each tuple representing a source.

    The process continues until the user decides not to add more sources or completes importing sources from a file. Invalid inputs at any stage are handled with error messages and re-prompting for valid input.

    Returns:
        list: A list of tuples, where each tuple contains:
            - Source name (str)
            - Right ascension (float)
            - Declination (float)
            - Variability rate (float or np.nan): 'np.nan' if the source is invariant or a float value representing the variability rate.

    Each tuple in the returned list represents an astronomical source with its respective coordinates and variability rate.
    """
    UserList = []

    while True:
        try:
            choice = str(input("Add sources to calculation? (yes/y or no/n): ")).lower()

            if choice in ['no', 'n']:
                return UserList
            elif choice in ['yes', 'y']:
                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")

    while True :
        open_file = str(input("Import data_src from a file? (yes/y or no/n): ")).lower()

        try:
            if open_file in ['yes', 'y']:
                print("Try : Catalog/exemple_src.txt")
                FILE_PATH = str(input("Enter a file_path : \n"))
                print("\n")
                print('-'*50)
                file_path = get_valid_file_path(FILE_PATH)
                print('-'*50)

                col1, ra, dec, value_var = np.loadtxt(file_path, unpack=True, usecols=(0, 1, 2, 3), dtype={'names': ('col1', 'ra', 'dec', 'valueVar'), 'formats': ('S25', 'f8', 'f8', 'f4')})
                name = [col1[data].decode().replace("_", " ") for data in range(len(col1))]

                for value in range(len(col1)):
                    UserList.append((name[value], ra[value], dec[value],value_var[value]))
                break

            elif open_file in ['no', 'n']:

                nbr_src = int(input("How many sources do you need to add ? "))
                item = 0
                print("\n")

                while item < nbr_src:
                    name = input('Enter source name : ')
                    ra = float(input('Enter right ascension : '))
                    dec = float(input('Enter declination : '))
                    value_var = input("Enter the value of variability rate of your object, enter nan if the object is invariant : ")
                    if value_var == 'nan':
                        value_var = np.nan
                    UserList.append((name, ra, dec, value_var))
                    item += 1
                    print(f"{colored(item, 'blue')} item added to the list \n")

                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")
   
    return UserList


def add_source_list(active_workflow) -> Table:
    """
    Prompt the user to add sources to a calculation and load a FITS file as a source list.

    Args:
        active_workflow (str): The path to the active workflow directory.

    Returns:
        Table: A table containing the source list data loaded from the FITS file. 
               If the user chooses not to add sources, an empty Table object is returned.

    Note:
        This function interacts with the user to determine if they wish to add sources to their calculation 
        by loading data from a FITS file. If the user opts to add sources, they are prompted to provide the 
        file path for the FITS file. The function then reads the specified FITS file and returns its contents 
        as an astropy Table object. The function ensures the file path is valid and handles any file-reading 
        errors. If the user decides not to add sources, the function returns an empty Table object.
    """
    
    print(f"You can add a {colored('.fits', 'blue')} file to your modeling ! ")
    while True:
        try:
            choice = str(input("Add sources to calculation? (yes/y or no/n): ")).lower()

            if choice in ['no', 'n']:
                add_source_table = Table()
                return add_source_table
            elif choice in ['yes', 'y']:
                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")
        
    path = str(input("Enter the file path : \n"))
    path = os.path.join(active_workflow, "catalog_data/add_sources.fits").replace("\\", "/")
    valid_path = get_valid_file_path(path)
    
    with fits.open(valid_path, memmap=True) as data:
        add_source_table = Table(data[1].data)
        
    return add_source_table
    

def get_coord_psr(name) -> SkyCoord:
    """
    Get the PSR coordinates from the SIMBAD database.

    Args:
        name (str): The name of the pulsar object.

    Returns:
        SkyCoord: A SkyCoord object representing the coordinates (RA and DEC) of the pulsar.
    """
    return SkyCoord(ra=Simbad.query_object(name)['RA'][0], dec=Simbad.query_object(name)['DEC'][0], unit=(u.hourangle, u.deg))


# --------------- None important function --------------- #


def py_to_xlsx(excel_data_path: str, count_rates: List, object_data: Dict, args: Tuple[str, str], radius: float) -> None:
    """
    Converts a list of count rates into an Excel file and saves it to the specified directory.

    Args:
        excel_data_path (str): Path to the directory where the Excel file will be saved.
        count_rates (List[float]): A list of count rates to be saved.
        object_data (Dict): Dictionary containing data about the observed astronomical object.
        args (Tuple[str, str]): Tuple of additional arguments, typically catalog identifiers 
                                (e.g., 'Xmm_DR13', 'CSC_2.0', 'Swift', 'eRosita', 'match').
        radius (float): Radius parameter related to the data, usually specifying a region of interest around the object.

    This function creates an Excel workbook, writes the count rates into it, and saves the file in the specified directory.
    The Excel file is named using a convention derived from the catalog type, radius, and object information. The file 
    name format is typically '{catalog}_{radius}_{object_name}.xlsx'.

    Returns:
        None

    Note:
        - The function supports various astronomical catalogs and adapts the file naming convention accordingly.
        - The count rates are written in a single column, starting from the first row of the Excel sheet.
        - The path and filename are processed to ensure compatibility with the file system.
    """

    
    if args[0] == "Xmm_DR13":
        cat = "xmm"
    elif args[0] == "CSC_2.0":
        cat = f"csc_{args[1]}"
    elif args[0] == "Swift":
        cat = "swi"
    elif args[0] == "eRosita":
        cat = "ero"
    elif args[0] == "match":
        cat = "xmmXchandra"
        
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    for item in range(len(count_rates)):
        sheet.cell(row=item + 1, column=1).value = count_rates[item]
    
    ct_rates_path = os.path.join(excel_data_path, f"{cat}_{radius}_{object_data['object_name']}.xlsx").replace("\\", "/") 
    wb.save(ct_rates_path.replace(" ", "_"))


def xlsx_to_py(excel_data_path: str, nearby_sources_table: Table, object_data: Dict, args: Tuple[str, str], radius: float) -> Tuple[List[float], Table]:
    """
    Converts Excel file data to Python format by reading count rate data and integrating it into an astropy Table.

    Args:
        excel_data_path (str): Path to the directory containing the Excel file.
        nearby_sources_table (Table): An astropy Table into which the count rates will be integrated.
        object_data (Dict): Dictionary containing data about the observed astronomical object.
        args (Tuple[str, str]): Tuple of additional arguments, typically catalog identifiers (e.g., 'Xmm_DR13', 'CSC_2.0').
        radius (float): Radius parameter related to the data, usually specifying a region of interest around the object.

    Returns:
        Tuple[List[float], Table]: A tuple where the first element is a list of count rates, 
                                    and the second element is the updated nearby sources table with a new 'count_rate' column.

    The function reads count rate data from the specified Excel file and integrates this data into the `nearby_sources_table`. 
    The Excel file is expected to be named based on catalog type and object information derived from `args` and `object_data`. 
    This function is designed to work with various astronomical catalogs, accommodating different data structures and formats.

    Note:
        - The function assumes the Excel file has a specific structure suited to astronomical count rate data.
        - The function handles the data reading and integration process, but error handling for file reading or data inconsistencies should be managed externally.
    """

    
    if args[0] == "Xmm_DR13":
        cat = "xmm"
    elif args[0] == "CSC_2.0":
        cat = f"csc_{args[1]}"
    elif args[0] == "Swift":
        cat = "swi"
    elif args[0] == "eRosita":
        cat = "ero"
    elif args[0] == "match":
        cat = "xmmXchandra"
        
    ct_rates_path = os.path.join(excel_data_path, f"{cat}_{radius}_{object_data['object_name']}.xlsx".replace(" ", "_"))
    wb = openpyxl.load_workbook(ct_rates_path)
    sheet = wb.active

    count_rates = []
    for item in range(len(nearby_sources_table)): 
        count_rates.append(sheet.cell(row = item + 1, column = 1).value)
        
    nearby_sources_table["count_rate"] = count_rates
    
    return count_rates, nearby_sources_table
