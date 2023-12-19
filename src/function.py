# --------------- Packages --------------- #

from astropy.io import fits
from astropy.table import Table
from astropy.time import Time
from astropy.coordinates import SkyCoord, Angle
from astropy import units as u
from termcolor import colored
from astroquery.simbad import Simbad
from typing import Dict, Tuple, Union, List
from scipy.optimize import curve_fit
from tqdm import tqdm
from jaxspec.data.util import fakeit_for_multiple_parameters

import sys
import os
import argparse
import numpy as np
import subprocess
import matplotlib.pyplot as plt
import openpyxl
import catalog_information as dict_cat
import catalog_class

# ---------------------------------------- #

def is_valid_file_path(file_path) -> bool:
    """
    Check if a file exists at the given file path.

    Parameters:
        file_path (str): The file path to be checked.

    Returns:
        bool: True if the file exists, False otherwise.

    Raises:
        None
    """
    try:
        if os.path.exists(file_path):
            return True
        else:
            return False
    except Exception as error:
        print(f"An error occured: {error}")


def get_valid_file_path(path) -> str:
    """
    Prompt the user for a valid file path until a valid one is provided.

    Parameters:
        path (str): The initial file path.

    Returns:
        str: A valid file path that exists.

    Raises:
        None
    """
    while True :
        if is_valid_file_path(path):
            print(f"The file at {colored(path, 'yellow')} is {colored('valid', 'green')}.")
            break
        else:
            print(f"The file at {colored(path, 'yellow')} doesn't exist or the path is {colored('invalid', 'red')}.")
            path = str(input("Enter the file path : \n"))
    return path


def choose_catalog(args_catalog) -> Tuple[str, str]:
    """
    Choose a catalog based on the provided keyword and return a valid file path for it.

    Parameters:
        Catalog (str): The catalog keyword, should be 'DR11' or 'DR13'.

    Returns:
        str: A valid file path for the selected catalog.

    Raises:
        argparse.ArgumentError: If an invalid catalog keyword is provided.
    """
    active_workflow = os.getcwd().replace("\\", "/")
    catalog_datapath = os.path.join(active_workflow, "data/catalog_data").replace("\\", "/")
    while True:
        try:
            if args_catalog == 'Xmm_DR13':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                xmm_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(xmm_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'CSC_2.0':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                print("-"*50)
                chandra_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                valid_path = get_valid_file_path(chandra_path)
                print((f"The file at {colored(valid_path, 'yellow')} is {colored('valid', 'green')}."))
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'Swift':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                swift_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(swift_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == 'eRosita':
                print("\n")
                print(f"{colored(args_catalog, 'yellow')} catalog is loading")
                eRosita_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                print("-"*50)
                valid_path = get_valid_file_path(eRosita_path)
                print("-"*50, "\n")
                return valid_path, args_catalog
            elif args_catalog == "compare_catalog":
                print("\n")
                print("Enter catalog keyword (Xmm_DR13/CSC_2.0/Swift/eRosita)")
                catalog_1 = str(input("First catalog : "))
                while True:
                    if catalog_1 == "Xmm_DR13":
                        catalog_1_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "CSC_2.0":
                        catalog_1_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "Swift":
                        catalog_1_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                        break
                    elif catalog_1 == "eRosita":
                        catalog_1_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                        break
                    else:
                        catalog_1 = str(input("Keyword unnaccepted, retry : "))
                valid_path_1 = get_valid_file_path(catalog_1_path)
                
                catalog_2 = str(input("Second catalog : "))
                while True:
                    if catalog_2 == "Xmm_DR13":
                        catalog_2_path = os.path.join(catalog_datapath, "4XMM_slim_DR13cat_v1.0.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "CSC_2.0":
                        catalog_2_path = os.path.join(catalog_datapath, "Chandra.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "Swift":
                        catalog_2_path = os.path.join(catalog_datapath, "Swift.fits").replace("\\", "/")
                        break
                    elif catalog_2 == "eRosita":
                        catalog_2_path = os.path.join(catalog_datapath, "eRosita.fits").replace("\\", "/")
                        break
                    else:
                        catalog_2 = str(input("Keyword unnaccepted, retry : "))
                valid_path_2 = get_valid_file_path(catalog_2_path)
                
                valid_path = (valid_path_1, valid_path_2, catalog_1, catalog_2)
                return valid_path, args_catalog
            elif args_catalog == "match":
                return 'matched_catalog', 'matched_catalog'
            else:
                raise argparse.ArgumentError(None, "invalid catalog keyword keyword. retry with Xmm_DR13, CSC_2.0, Swift, eRosita")
        except argparse.ArgumentError as error:
            print(f'An error occured : {error}')
            args_catalog = str(input("Enter a new key word : \n"))


def define_sources_list() -> Table:
    """
    Prompts the user to define a list of sources for calculations. This function allows users to add sources either manually or by importing data from a file.

    Returns:
    SRC_LIST (list): A list of source tuples, each containing the source name, right ascension (ra), and declination (dec).

    The function begins by asking the user whether they want to add sources to the calculation.
    If the user chooses to add sources manually, they will be prompted to enter the source details (name, ra, and dec) for each source.
    If the user chooses to import data from a file, the function reads data from the specified file and extracts the necessary columns.
    The extracted data is then added to the SRC_LIST as source tuples.

    The function continues to prompt the user until they have either manually added all the sources they need or imported sources from a file.
    If the user enters invalid input at any point, the function will display an error message and continue to prompt for valid input.
    
    """
    UserList = []

    while True:
        try:
            choice = str(input("Add sources to calculation? (yes/y or no/n): ")).lower()

            if choice in ['no', 'n']:
                return UserList
            elif choice in ['yes', 'y']:
                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")

    while True :
        open_file = str(input("Import data_src from a file? (yes/y or no/n): ")).lower()

        try:
            if open_file in ['yes', 'y']:
                print("Try : Catalog/exemple_src.txt")
                FILE_PATH = str(input("Enter a file_path : \n"))
                print("\n")
                print('-'*50)
                file_path = get_valid_file_path(FILE_PATH)
                print('-'*50)

                col1, ra, dec, value_var = np.loadtxt(file_path, unpack=True, usecols=(0, 1, 2, 3), dtype={'names': ('col1', 'ra', 'dec', 'valueVar'), 'formats': ('S25', 'f8', 'f8', 'f4')})
                name = [col1[data].decode().replace("_", " ") for data in range(len(col1))]

                for value in range(len(col1)):
                    UserList.append((name[value], ra[value], dec[value],value_var[value]))
                break

            elif open_file in ['no', 'n']:

                nbr_src = int(input("How many sources do you need to add ? "))
                item = 0
                print("\n")

                while item < nbr_src:
                    name = input('Enter source name : ')
                    ra = float(input('Enter right ascension : '))
                    dec = float(input('Enter declination : '))
                    value_var = input("Enter the value of variability rate of your object, enter nan if the object is invariant : ")
                    if value_var == 'nan':
                        value_var = np.nan
                    UserList.append((name, ra, dec, value_var))
                    item += 1
                    print(f"{colored(item, 'blue')} item added to the list \n")

                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")
   
    return UserList


def add_source_list(active_workflow) -> Table:
    """
    Prompt the user to add sources to a calculation and load a FITS file as a source list.

    Args:
        active_workflow (str): The path to the active workflow directory.

    Returns:
        Table: A table containing the source list data loaded from the FITS file.

    This function prompts the user to add sources to a calculation and allows them to load a FITS file as a source list.
    If the user chooses to add sources, they will be prompted to enter the file path of the FITS file.
    The function then opens the FITS file, reads the data, and returns it as a Table object.

    If the user chooses not to add sources, an empty Table is returned.
    """
    
    print(f"You can add a {colored('.fits', 'blue')} file to your modeling ! ")
    while True:
        try:
            choice = str(input("Add sources to calculation? (yes/y or no/n): ")).lower()

            if choice in ['no', 'n']:
                add_source_table = Table()
                return add_source_table
            elif choice in ['yes', 'y']:
                break
            else:
                raise ValueError("Invalid input. Please enter 'yes' or 'y' for yes, 'no' or 'n' for no.")
        except ValueError as error:
            print(f"An error occured {error}")
        
    path = str(input("Enter the file path : \n"))
    path = os.path.join(active_workflow, "catalog_data/add_sources.fits").replace("\\", "/")
    valid_path = get_valid_file_path(path)
    
    with fits.open(valid_path, memmap=True) as data:
        add_source_table = Table(data[1].data)
        
    return add_source_table
    

def get_coord_psr(name) -> SkyCoord:
    """
    Get the PSR coordinates from the SIMBAD database.

    Parameters:
    name (str): The name of the pulsar object.

    Returns:
    SkyCoord: A SkyCoord object representing the coordinates (RA and DEC) of the pulsar.
    """
    return SkyCoord(ra=Simbad.query_object(name)['RA'][0], dec=Simbad.query_object(name)['DEC'][0], unit=(u.hourangle, u.deg))


def scaled_ct_rate(D, OptCtRate, effareaX, effareaY) -> float:
    """
    Scale a given count rate based on an angular distance and effective area.

    Parameters:
    D (float): The angular distance.
    OptCtRate (float): The original count rate.
    effareaX (array-like): Effective area data points (X values).
    effareaY (array-like): Effective area data points (Y values).

    Returns:
    float: The scaled count rate.
    """
    return OptCtRate * np.interp(D,effareaX,effareaY)


def ang_separation(reference, obj) -> Angle:
    """
    Calculate the angular separation between two celestial objects.

    Parameters:
    reference (SkyCoord): The reference object's coordinates.
    obj (SkyCoord): The coordinates of the object to which the separation is calculated.

    Returns:
    Quantity: The angular separation between the two objects.
    """
    return reference.separation(obj)


def signal_to_noise(SrcCtsRate, BkgSrcRates, InstBkgd, ExpTime) -> float:
    """
    Calculate the signal-to-noise ratio (S/N) given various parameters.

    Parameters:
    SrcCtsRate (float): Source count rate.
    BkgSrcRates (array-like): Count rates of background sources.
    InstBkgd (float): Instrumental and particle background count rate.
    ExpTime (float): Exposure time.

    Returns:
    float: The signal-to-noise ratio (S/N).
    """
    SNR = (SrcCtsRate*ExpTime) / np.sqrt(ExpTime*(SrcCtsRate+np.sum(BkgSrcRates)+InstBkgd))
    return SNR


def nominal_pointing_info(simulation_data, NearbySRCposition) -> None:
    """
    Calculate and print various information related to nominal pointing.

    Parameters:
    SIM_parameters (dict): Dictionary containing simulation parameters.
    NearbySRCposition (SkyCoord): Coordinates of nearby sources.

    Returns:
    None
    """

    object_data = simulation_data['object_data']
    telescop_data = simulation_data['telescop_data']
    
    SRCnominalDIST = ang_separation(NearbySRCposition, SkyCoord(ra=object_data['object_position'].ra, dec=object_data['object_position'].dec)).arcmin
    SRCscaleRates = scaled_ct_rate(SRCnominalDIST, simulation_data["nearby_sources_table"]["count_rate"], telescop_data["EffArea"], telescop_data["OffAxisAngle"])
    PSRcountrates = object_data['count_rate']

    print('PSR S/N at Nominal Pointing ' + str(signal_to_noise(PSRcountrates, SRCscaleRates, simulation_data["INSTbkgd"], simulation_data["EXPtime"])))
    print("PSR count rate at Nominal pointing = " + str(PSRcountrates) + "cts/sec")
    print("BKG sources count rate at Nominal Pointing = " + str(np.sum(SRCscaleRates)) + "cts/sec")
    print("             Individual BKG Sources rates:")
    print(str(SRCscaleRates))
    print("             BKG sources distance from PSR (\')")
    print(SRCnominalDIST)
    print("--------------------------------------------------")


def calculate_opti_point(simulation_data, nearby_src_position) -> Tuple[int, float, float, dict]:
    """
    Calculate the optimal pointing position for a telescope to maximize the signal-to-noise ratio (SNR).

    Args:
        simulation_data (dict): A dictionary containing simulation data including telescope data and object data.
        nearby_src_position (numpy.ndarray): An array containing the positions of nearby sources.

    Returns:
        Tuple[int, float, float, dict]: A tuple containing:
            - OptimalPointingIdx (int): Index of the optimal pointing position.
            - SRCoptimalSEPAR (float): Angular separation of nearby sources from the optimal pointing position (in arcminutes).
            - SRCoptimalRATES (float): Scaled count rates of nearby sources at the optimal pointing position.
            - vector_dictionary (dict): A dictionary containing various vectors and results, including:
                - 'SampleRA': Array of right ascensions for sample pointing positions (in degrees).
                - 'SampleDEC': Array of declinations for sample pointing positions (in degrees).
                - 'PSRrates': Array of scaled count rates for the target object at sample pointing positions.
                - 'SRCrates': Array of total scaled count rates for nearby sources at sample pointing positions.
                - 'SNR': Array of signal-to-noise ratios at sample pointing positions.
    """
    
    min_value, max_value, step = -7.0, 7.1, 0.05
    DeltaRA = Angle(np.arange(min_value, max_value, step), unit=u.deg)/60
    DeltaDEC = Angle(np.arange(min_value, max_value, step), unit=u.deg)/60
    
    telescop_data = simulation_data["telescop_data"]
    object_data = simulation_data["object_data"]
    nearby_sources_table = simulation_data['nearby_sources_table']
    
    RA_grid, DEC_grid = np.meshgrid(DeltaRA, DeltaDEC)

    SampleRA = object_data["object_position"].ra.deg + RA_grid.flatten().deg
    SampleDEC = object_data["object_position"].dec.deg + DEC_grid.flatten().deg

    NICERpointing = SkyCoord(ra=SampleRA*u.deg, dec=SampleDEC*u.deg)

    PSRseparation = ang_separation(object_data["object_position"], NICERpointing).arcmin
    nearby_src_position = nearby_src_position.reshape(1, -1)
    NICERpointing = NICERpointing.reshape(-1, 1)
    SRCseparation = ang_separation(nearby_src_position, NICERpointing).arcmin

    PSRcountrateScaled = scaled_ct_rate(PSRseparation, object_data['count_rate'], telescop_data["EffArea"], telescop_data["OffAxisAngle"])

    count_rate = nearby_sources_table['count_rate']
    SRCcountrateScaled = scaled_ct_rate(SRCseparation, count_rate, telescop_data["EffArea"], telescop_data["OffAxisAngle"])

    SNR, PSRrates, SRCrates  = np.zeros((3, len(DeltaRA) * len(DeltaDEC)))
    for item in range(len(PSRcountrateScaled)):
        PSRrates[item] = PSRcountrateScaled[item]
        SRCrates[item] = np.sum(SRCcountrateScaled[item])
        SNR[item] = signal_to_noise(PSRrates[item], SRCrates[item], simulation_data["INSTbkgd"], simulation_data["EXPtime"])

    OptimalPointingIdx = np.where(SNR==max(SNR))[0][0]

    SRCoptimalSEPAR = ang_separation(nearby_src_position, SkyCoord(ra=SampleRA[OptimalPointingIdx]*u.degree, dec=SampleDEC[OptimalPointingIdx]*u.degree)).arcmin
    SRCoptimalRATES = scaled_ct_rate(SRCoptimalSEPAR, nearby_sources_table["count_rate"], telescop_data["EffArea"], telescop_data["OffAxisAngle"])

    vector_dictionary = {
        'SampleRA': SampleRA,
        'SampleDEC': SampleDEC,
        'PSRrates': PSRrates,
        'SRCrates': SRCrates,
        'SNR': SNR
    }

    return OptimalPointingIdx, SRCoptimalSEPAR, SRCoptimalRATES, vector_dictionary


def optimal_point_infos(vector_dictionary, OptimalPointingIdx, SRCoptimalRATES) -> None:
    """
    Print information for the optimal NICER pointing that maximizes the signal-to-noise ratio (S/N).

    This function prints information about the S/N ratio, pulsar count rate, background sources count rate,
    individual background source rates, and the optimal pointing coordinates for the NICER telescope.

    Args:
        vector_dictionary (dict): A dictionary contai
        # if isinstance(model_value, tuple):
        #     model_value = model_value[0]ning result vectors, including sampled RA and DEC positions,
                                  pulsar count rates, SRC count rates, and the S/N ratio for each pointing.
        OptimalPointingIdx (int): The index of the optimal pointing in the result vectors.
        SRCoptimalRATES (float): The SRC count rate at the optimal pointing.
    """

    # Print info for the optimal NICER pointing that maximizes the S/N ratio
    print ("PSR S/N at Optimal Pointing " + str(vector_dictionary['SNR'][OptimalPointingIdx]))
    print ("PSR count rate at Optimal pointing = " + str(vector_dictionary["PSRrates"][OptimalPointingIdx]) + " cts/sec")
    print ("BKG sources count rate at Optimal pointing = " + str(vector_dictionary["SRCrates"][OptimalPointingIdx]) + " cts/sec")
    print ("     Individual BKG Sources: " )
    print (str(SRCoptimalRATES))
    #print "     Distance from Optimal Pointing (\")"
    #print str(SRCoptimalSEPAR)
    print ("Optimal Pointing:  " + str(vector_dictionary["SampleRA"][OptimalPointingIdx]) + "  " + str(vector_dictionary["SampleDEC"][OptimalPointingIdx]))
    print ("----------------------------------------------------------------------")


def data_map(simulation_data, vector_dictionary, OptimalPointingIdx, NearbySRCposition) -> None:
    """
    Plot the map of the Signal-to-Noise (S/N) ratio as a function of NICER pointing.

    Parameters:
    - SIM_parameters (dict): A dictionary containing simulation parameters, including the pulsar position.
    - Vector_Dictionary (dict): A dictionary containing vector data, including SampleRA, SampleDEC, and SNR.
    - OptimalPointingIdx (int): The index of the optimal pointing in Vector_Dictionary.
    - NearbySRCposition (SkyCoord): SkyCoord object representing the positions of nearby sources.

    Returns:
    None
    
    """
    os_dictionary = simulation_data["os_dictionary"]
    object_data = simulation_data['object_data']

    # Plot the map of S/N ratio as function of NICER pointing
    ra_opti = vector_dictionary['SampleRA'][OptimalPointingIdx]
    dec_opti = vector_dictionary['SampleDEC'][OptimalPointingIdx]
    
    nearby_ra = [NearbySRCposition[item].ra.value for item in range(len(NearbySRCposition))]
    nearby_dec = [NearbySRCposition[item].dec.value for item in range(len(NearbySRCposition))]
    
    figure, axes = plt.subplots(1, 1, figsize=(15, 8))
    figure.suptitle(f"S/N map for {object_data['object_name']}\nOptimal pointing point : {ra_opti} deg, {dec_opti} deg")
    
    axes.invert_xaxis()
    sc = axes.scatter(vector_dictionary['SampleRA'], vector_dictionary['SampleDEC'], c=vector_dictionary["SNR"], s=10, edgecolor='face')
    axes.scatter(nearby_ra, nearby_dec, marker='.', color='black', label=f"Nearby sources : {len(nearby_ra)}")
    axes.scatter(object_data["object_position"].ra, object_data["object_position"].dec, marker='*', color='green', label=f"{object_data['object_name']}")
    axes.scatter(ra_opti, dec_opti, s=50, marker='+', color='red', label="Optimal Pointing Point")
    
    axes.set_xlabel('Right Ascension [deg]', fontsize='large')
    axes.set_ylabel('Declination [deg]', fontsize='large')
    axes.legend(loc="upper right", ncol=2)
    cbar = figure.colorbar(sc, ax=axes)
    cbar.set_label('S/N')
    
    key = simulation_data["os_dictionary"]["catalog_key"]
    name = object_data["object_name"]
    plt.savefig(os.path.join(os_dictionary['img'], f"{key}_SNR_{name}.png".replace(" ", "_")))
    plt.show()


def count_rates(nearby_src_table, model_dictionary, telescop_data) -> Tuple[List[float], Table]:
    """
    Calculate X-ray count rates for nearby sources using PIMMS modeling.

    This function calculates X-ray count rates for a set of nearby sources based on their model information and associated parameters.
    It uses PIMMS (Portable, Interactive Multi-Mission Simulator) to perform the modeling and computes the count rates.

    Parameters:
        nearby_src_table (Table): A table containing data on nearby sources, including model information, X-ray flux, and column density.
        model_dictionary (dict): A dictionary mapping sources (e.g., "src_0") to model and parameter information.

    Returns:
        tuple: A tuple containing two elements:
            - A NumPy array of X-ray count rates for each nearby source.
            - An updated 'nearby_src_table' with the added 'Count Rates' column.

    Note:
    - PIMMS modeling commands are generated for each source based on the model, model value, X-ray flux, and column density.
    - The 'Count Rates' column is added to the 'nearby_src_table' with the calculated count rates.
    """
    number_source = len(model_dictionary)
    count_rates = np.array([], dtype=float)
    
    telescop_name = telescop_data['telescop_name']
    min_value = telescop_data['min_value']
    max_value = telescop_data['max_value']
    energy_band = telescop_data['energy_band']
    
    for item in range(number_source):
        model = model_dictionary[f"src_{item}"]["model"]
        model_value = model_dictionary[f"src_{item}"]["model_value"]
        xmm_flux =  model_dictionary[f"src_{item}"]["flux"]
        nh_value = model_dictionary[f"src_{item}"]["column_dentsity"]
                
        pimms_cmds = f"instrument {telescop_name} {min_value}-{max_value}\nfrom flux ERGS {energy_band}\nmodel galactic nh {nh_value}\nmodel {model} {model_value} 0.0\ngo {xmm_flux}\nexit\n"
        
        with open('pimms_script.xco', 'w') as file:
            file.write(pimms_cmds)
            file.close()

        result = subprocess.run(['pimms', '@pimms_script.xco'], stdout=subprocess.PIPE).stdout.decode(sys.stdout.encoding)
        count_rate = float(result.split("predicts")[1].split('cps')[0])
        count_rates = np.append(count_rates, count_rate)
        
    nearby_src_table["count_rate"] = count_rates
        
    return count_rates, nearby_src_table

# --------------- multiple sources catalog to unique sources catalog --------------- #

def unique_dict(name_list: List) -> Dict:
    """
    Create a dictionary that associates names with their indices in a list.

    Args:
        name_list (List): A list of names.

    Returns:
        Dict: A dictionary where keys are names and values are lists of corresponding indices.
    """
    index_dict = {}
    duplicate_dict = {}
    
    for index, item in enumerate(name_list):
        if item in index_dict:
            if item in duplicate_dict:
                duplicate_dict[item].append(index)
            else:
                duplicate_dict[item] = [index_dict[item], index]
        else:
            index_dict[item] = index
    return duplicate_dict


def insert_row(unique_sources_dict: Dict, new_row: List[Tuple]) -> Dict:
    """
    Insert a new row into a dictionary and maintain sorted order based on values.

    Args:
        unique_sources_dict (Dict): A dictionary with names as keys and lists of indices as values.
        new_row (List[Tuple]): A list of tuples containing (name, index) to be inserted.

    Returns:
        Dict: Updated dictionary with the new row inserted.
    """
    
    new_row.sort(key=lambda x: x[1])
    
    for key, value in new_row:
        # Convertir le dictionnaire actuel en liste de paires clé-valeur
        items = list(unique_sources_dict.items())
        # Trouver l'emplacement approprié pour la nouvelle paire
        for index, (_, liste_valeurs) in enumerate(items):
            if liste_valeurs[0] > value:
                # Insérer la nouvelle paire avant cette position
                items.insert(index, (key, [value]))
                break
        else:
            # Si aucune valeur plus grande n'a été trouvée, ajouter à la fin
            items.append((key, [value]))
        # Recréer le dictionnaire
        unique_sources_dict = dict(items)
    return unique_sources_dict  


def replace_nan_value(key: str, unique_table: Table) -> Table:
    """
    Replace NaN values in a table's specified columns with their minimum non-NaN values.

    Args:
        key (str): The catalog key to determine which columns to process.
        unique_table (Table): A table containing data with NaN values.

    Returns:
        Table: Updated table with NaN values replaced.
    """
    
    flux_obs = dict_cat.dictionary_catalog[key]["flux_obs"]
    flux_obs_err = dict_cat.dictionary_catalog[key]["flux_obs_err"]
    band_flux_obs = dict_cat.dictionary_catalog[key]["band_flux_obs"]
    band_flux_obs_err = dict_cat.dictionary_catalog[key]["band_flux_obs_err"]

    flux_list = [flux_obs, flux_obs_err, band_flux_obs, band_flux_obs_err[0], band_flux_obs_err[1]]

    flux_name = []

    for flux in flux_list:
        if isinstance(flux, str):
            flux_name.append(flux)
        else:
            for item in range(len(flux)):
                flux_name.append(flux[item])

    for name in flux_name:
        flux_data = []
        index_data = []
        for index, flux in enumerate(unique_table[name]):
            if not np.isnan(flux):
                flux_data.append(flux)
            else:
                index_data.append(index)
        min_value = np.min(flux_data)
        
        for index in index_data:
            unique_table[name][index] = min_value
            
    return unique_table


def create_unique_sources_catalog(nearby_sources_table: Table, column_name: List) -> Table:
    """
    Create a unique sources catalog based on a nearby sources table and catalog-specific column names.

    Args:
        nearby_sources_table (Table): A table containing nearby sources data.
        column_name (List): A list of column names used for catalog-specific data.

    Returns:
        Table: A table representing the unique sources catalog.
    """
    
    key = column_name["catalog_name"]

    dict_flux_name = {"flux_obs": dict_cat.dictionary_catalog[key]["flux_obs"],
                        "flux_obs_err": dict_cat.dictionary_catalog[key]["flux_obs_err"],
                        "band_flux_obs": dict_cat.dictionary_catalog[key]["band_flux_obs"],
                        "band_flux_obs_err": dict_cat.dictionary_catalog[key]["band_flux_obs_err"]}
    
    list_flux_name = [dict_flux_name["flux_obs"], dict_flux_name["flux_obs_err"], dict_flux_name["band_flux_obs"], dict_flux_name["band_flux_obs_err"][0], dict_flux_name["band_flux_obs_err"][1]]
    
    flux_name = []
    for value in list_flux_name:
        if isinstance(value, str):
            flux_name.append(value)
        else:
            for item in value:
                flux_name.append(item)
                
    for flux in flux_name:
        min_value = np.nanmean(nearby_sources_table[flux])
        nan_mask = np.isnan(nearby_sources_table[flux])
        nearby_sources_table[flux][nan_mask] = min_value

    unique_sources_dict = unique_dict(nearby_sources_table[column_name["source_name"]])
    
    new_row = []
    for index, name in enumerate(nearby_sources_table[column_name["source_name"]]):
        if name not in unique_sources_dict.keys():
            new_row.append((name, index))
        
    sources_dict = insert_row(unique_sources_dict=unique_sources_dict, new_row=new_row)
    
    
    if key == "Chandra":
        
        iauname_col, ra_col, dec_col = [], [], []
        for key, value in list(sources_dict.items()):
            iauname_col.append(key)
            ra_col.append(np.mean([nearby_sources_table[column_name["right_ascension"]][index] for index in value]))
            dec_col.append(np.mean([nearby_sources_table[column_name["declination"]][index] for index in value]))
        
        unique_table = Table()
        unique_table["Chandra_IAUNAME"] = iauname_col
        unique_table["RA"] = ra_col
        unique_table["DEC"] = dec_col
        
    if key == "Swift":
        
        iauname_col, ra_col, dec_col = [], [], []
        for key, value in list(sources_dict.items()):
            iauname_col.append(key)
            ra_col.append(np.mean([nearby_sources_table[column_name["right_ascension"]][index] for index in value]))
            dec_col.append(np.mean([nearby_sources_table[column_name["declination"]][index] for index in value]))
        
        unique_table = Table()
        unique_table["Swift_IAUNAME"] = iauname_col
        unique_table["RA"] = ra_col
        unique_table["DEC"] = dec_col
        
    if key == "eRosita":
    
        iauname_col, ra_col, dec_col = [], [], []
        for key, value in list(sources_dict.items()):
            iauname_col.append(key)
            ra_col.append(np.mean([nearby_sources_table[column_name["right_ascension"]][index] for index in value]))
            dec_col.append(np.mean([nearby_sources_table[column_name["declination"]][index] for index in value]))
        
        unique_table = Table()
        unique_table["eRosita_IAUNAME"] = iauname_col
        unique_table["RA"] = ra_col
        unique_table["DEC"] = dec_col
        
    for flux in flux_name:
        data = []
        for value in list(sources_dict.values()):
            if len(value) != 1:
                new_value = np.mean([nearby_sources_table[flux][index] for index in value])
            else:
                new_value = nearby_sources_table[flux][value[0]]
            data.append(new_value)
        unique_table[flux] = data
        
    return unique_table


# ---------------------------------------------------------------------------------- #


def vignetting_factor(OptimalPointingIdx, vector_dictionary, simulation_data, data, nearby_sources_table) -> Tuple[List[float], Table]:
    """
    Calculate the vignetting factors for nearby sources and the target object based on their distances.

    Args:
        OptimalPointingIdx (int): Index of the optimal pointing position.
        vector_dictionary (dict): A dictionary containing vectors for RA and DEC.
        simulation_data (dict): A dictionary containing simulation data including object data, telescope data, etc.
        data (tuple): A tuple containing RA, DEC, and name data.
        nearby_sources_table (Table): A table containing information about nearby sources.

    Returns:
        Tuple[List[float], Table]: A tuple containing the calculated vignetting factors and an updated nearby sources table.
    """
    
    ra, dec, name = data
    
    object_data = simulation_data["object_data"]
    EffArea, OffAxisAngle = simulation_data["telescop_data"]["EffArea"], simulation_data["telescop_data"]["OffAxisAngle"]
    
    optipoint_ra, optipoint_dec = vector_dictionary['SampleRA'][OptimalPointingIdx], vector_dictionary['SampleDEC'][OptimalPointingIdx]
    
    def calculate_vignetting_factor(D, effareaX, effareaY):
        return np.interp(D,effareaX,effareaY)
    
    vignetting_factor, distance = np.array([], dtype=float), np.array([], dtype=float)
    
    for index in range(len(nearby_sources_table)):
        SRCposition  = SkyCoord(ra=nearby_sources_table[ra][index]*u.degree, dec=nearby_sources_table[dec][index]*u.degree)
        SRCnominalDIST = ang_separation(SRCposition, SkyCoord(ra=optipoint_ra, dec=optipoint_dec, unit=u.deg)).arcmin
        distance = np.append(distance, SRCnominalDIST)
        vignetting = calculate_vignetting_factor(SRCnominalDIST, EffArea, OffAxisAngle)
        vignetting_factor = np.append(vignetting_factor, vignetting)
    
    optimal_pointing_point = SkyCoord(ra=optipoint_ra, dec=optipoint_dec, unit=u.deg)
    psr_position = SkyCoord(ra=object_data['object_position'].ra, dec=object_data['object_position'].dec, unit=u.deg)
    distance_psr_to_optipoint = ang_separation(psr_position, optimal_pointing_point).arcmin
    vignetting_factor_psr2optipoint = calculate_vignetting_factor(distance_psr_to_optipoint, EffArea, OffAxisAngle)

    max_vignet, min_distance  = np.max(vignetting_factor), np.min(distance)
    max_vignet_index, min_distance_index = np.argmax(vignetting_factor), np.argmin(distance)
    
    print(f"\nThe closest source of the optimal pointing point is : {colored(nearby_sources_table[name][min_distance_index], 'magenta')}.")
    print(f"The distance between {colored(nearby_sources_table[name][min_distance_index], 'yellow')} and optimal pointing point is {colored(min_distance, 'blue')} arcmin.\n"
          f"With a vignetting factor of : {colored(max_vignet, 'light_green')} ")
    print(f"The distance between {colored(object_data['object_name'], 'yellow')} and optimal pointing point is {colored(distance_psr_to_optipoint, 'blue')} arcmin,\n"
          f"with a vagnetting factor of : {colored(vignetting_factor_psr2optipoint, 'light_green')}")
    
    nearby_sources_table["vignetting_factor"] = vignetting_factor
    
    return vignetting_factor, nearby_sources_table


def write_fits_file(nearby_sources_table, simulation_data) -> None:
    """
    Write the nearby sources table to a FITS file and open it with TOPCAT.

    Args:
        nearby_sources_table (Table): A table containing information about nearby sources.
        simulation_data (dict): A dictionary containing simulation data and file path information.

    Returns:
        None
    """
    
    try:
        os_dictionary = simulation_data["os_dictionary"]
        key = os_dictionary["catalog_key"]
        cloesest_dataset_path = os_dictionary["cloesest_dataset_path"]
        nearby_sources_table_path = os.path.join(cloesest_dataset_path, f"{key}_nearby_sources_table.fits").replace("\\", "/")
        nearby_sources_table.write(nearby_sources_table_path, format='fits', overwrite=True)
        print(f"Nearby sources table was created in : {colored(nearby_sources_table_path, 'magenta')}")
        
        topcat_path = os_dictionary["topcat_software_path"]
        command = f"java -jar {topcat_path} {nearby_sources_table_path}"
        subprocess.run(command)
        
    except Exception as error:
        print(f"{colored('An error occured : ', 'red')} {error}")
    
    
def modeling(vignetting_factor: List, simulation_data: Dict, column_dictionary: Dict, catalog_name: str) -> None:
    """
    Perform modeling of nearby sources using a power-law model and create a plot.

    Args:
        vignetting_factor (List): List of calculated vignetting factors for nearby sources.
        simulation_data (Dict): A dictionary containing simulation data and object information.
        column_dictionary (Dict): A dictionary containing column names for flux and energy band.
        catalog_name (str): Name of the catalog used for modeling.

    Returns:
        None
    """
    
    object_data = simulation_data["object_data"]
    os_dictionary = simulation_data["os_dictionary"]
    nearby_sources_table = simulation_data["nearby_sources_table"]
    number_source = len(nearby_sources_table)
    
    flux_name, err_flux_name, energy_band = column_dictionary["band_flux_obs"], column_dictionary["band_flux_obs_err"], column_dictionary["energy_band"]
    photon_index_list, constant_list = np.array([], dtype=float), np.array([], dtype=float)
    main_flux_obs, main_err_flux_obs = [], []
    
    

    def power_law(vignetting_factor, energy_band, constant, gamma):
        sigma = column_dictionary["sigma"]
        return (constant * energy_band ** (-gamma) * np.exp(-sigma * 3e20)) * vignetting_factor
    
    
    def sum_power_law(energy_band, constant, photon_index):
        summed_power_law = 0.0
        for cst, pho in zip(constant, photon_index):
            summed_power_law += (cst * energy_band ** (-pho))
        return summed_power_law

    
    for item in range(number_source):
        flux_obs = [nearby_sources_table[name][item] for name in flux_name]
        main_flux_obs.append(flux_obs)
        
        err_flux_obs = [nearby_sources_table[err_name][item] for err_name in err_flux_name]
        main_err_flux_obs.append(err_flux_obs)
        try:
            popt, pcov = curve_fit(lambda energy_band, constant, gamma: power_law(vignetting_factor[item], energy_band, constant, gamma), energy_band, flux_obs, sigma=err_flux_obs)
            constant, photon_index = popt
        except Exception as error:
            constant = 1e-14
            photon_index = 1.7
        photon_index_list = np.append(photon_index_list, photon_index)
        constant_list = np.append(constant_list, constant)
        
        
    flux_obs_array = []
    for index in range(number_source):
        flux_obs_array.append(power_law(vignetting_factor[index], energy_band, constant_list[index], photon_index_list[index]))
        
    percentiles = np.percentile(flux_obs_array, (16, 50, 84), axis=0)
    
    nrows, ncols = 1, 3
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(15, 6), sharex=True)
    fig.suptitle(f"Modeling nearby sources with {catalog_name}\n{object_data['object_name']}", fontsize=16)
    fig.text(0.5, 0.04, 'Energy [keV]', ha='center', va='center', fontsize=12)
    fig.text(0.07, 0.5, 'Flux [erg/cm2/s]', ha='center', va='center', rotation='vertical', fontsize=12)

    ax00, ax01, ax02 = axes[0], axes[1], axes[2]
    ax00.set_title("All nearby sources power law")
    ax00.set_xscale("log")

    for index in range(number_source):
        ax00.plot(energy_band, power_law(vignetting_factor[index], energy_band, constant_list[index], photon_index_list[index]))
    
    ax01.plot(energy_band, percentiles[1], color="navy", label="mean value")
    ax01.plot(energy_band, percentiles[0], color="royalblue", ls="--", linewidth=1.5, label="$16^{th}$ percentile")
    ax01.plot(energy_band, percentiles[2], color="midnightblue", ls="--", linewidth=1.5, label="$84^{th}$ percentile")
    ax01.fill_between(energy_band, percentiles[0], percentiles[2], alpha=0.3, color="navy", hatch='\\'*3, label="envelop")
    ax01.legend(loc="upper right", fontsize=8, ncols=2)

    ax02.plot(energy_band, sum_power_law(energy_band, constant_list, photon_index_list), color="darkmagenta", ls='-.', label="summed power law")
    ax02.legend(loc="upper right", fontsize=10)
    
    key = simulation_data["os_dictionary"]["catalog_key"]
    name = object_data['object_name']
    plt.savefig(os.path.join(os_dictionary["img"], f"{key}_modeling_{name}.png".replace(" ", "_")))
    plt.show()
    
    
# --------------- None important function --------------- #


def py_to_xlsx(excel_data_path: str, count_rates: List, object_data: Dict, args: Tuple[str, str], radius: float) -> None:
    """
    Converts and saves Python data into an Excel file.

    Args:
    excel_data_path (str): The path to the directory where the Excel file will be saved.
    count_rates (List): A list of count rates to be saved.
    object_data (Dict): A dictionary containing data about the observed object.
    args (Tuple[str, str]): A tuple containing additional arguments, typically catalog identifiers.
    radius (float): The radius parameter related to the data.

    This function creates an Excel workbook and writes the provided count rates into it. The file is named based on 
    the provided arguments and object data, and is saved in the specified directory.

    Note:
    - The file naming convention is derived from the catalog type and object information.
    - The function currently supports different catalogs (e.g., Xmm_DR13, CSC_2.0, Swift, eRosita, match).
    """

    
    if args[0] == "Xmm_DR13":
        cat = "xmm"
    elif args[0] == "CSC_2.0":
        cat = f"csc_{args[1]}"
    elif args[0] == "Swift":
        cat = "swi"
    elif args[0] == "eRosita":
        cat = "ero"
    elif args[0] == "match":
        cat = "xmmXchandra"
        
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    for item in range(len(count_rates)):
        sheet.cell(row=item + 1, column=1).value = count_rates[item]
    
    ct_rates_path = os.path.join(excel_data_path, f"{cat}_{radius}_{object_data['object_name']}.xlsx").replace("\\", "/") 
    wb.save(ct_rates_path.replace(" ", "_"))


def xlsx_to_py(excel_data_path: str, nearby_sources_table: Table, object_data: Dict, args: Tuple[str, str], radius: float) -> Tuple[List[float], Table]:
    """
    Reads count rate data from an Excel file and integrates it into a Python table.

    Args:
    excel_data_path (str): The path to the directory containing the Excel file.
    nearby_sources_table (Table): The table into which the count rates will be integrated.
    object_data (Dict): A dictionary containing data about the observed object.
    args (Tuple[str, str]): A tuple containing additional arguments, typically catalog identifiers.
    radius (float): The radius parameter related to the data.

    The function opens the specified Excel workbook, reads count rate data from it, and adds these count rates to the 
    provided table under a new column 'count_rate'.

    Returns:
    Tuple[List[float], Table]: A tuple containing the list of count rates and the updated nearby sources table.

    Note:
    - The file to be read is named based on the catalog type and object information.
    - The function supports different catalogs (e.g., Xmm_DR13, CSC_2.0, Swift, eRosita, match).
    """

    
    if args[0] == "Xmm_DR13":
        cat = "xmm"
    elif args[0] == "CSC_2.0":
        cat = f"csc_{args[1]}"
    elif args[0] == "Swift":
        cat = "swi"
    elif args[0] == "eRosita":
        cat = "ero"
    elif args[0] == "match":
        cat = "xmmXchandra"
        
    ct_rates_path = os.path.join(excel_data_path, f"{cat}_{radius}_{object_data['object_name']}.xlsx".replace(" ", "_"))
    wb = openpyxl.load_workbook(ct_rates_path)
    sheet = wb.active

    count_rates = []
    for item in range(len(nearby_sources_table)): 
        count_rates.append(sheet.cell(row = item + 1, column = 1).value)
        
    nearby_sources_table["count_rate"] = count_rates
    
    return count_rates, nearby_sources_table


# --------------- Software function ---------- # 


def load_relevant_sources(cat: str, file_to_load: str) -> Dict:
    """
    Loads and processes source data from a specified catalog and returns it in a structured format.

    Args:
    cat (str): Name of the catalog to load (e.g., "Swift", "XMM").
    file_to_load (str): Path to the FITS file containing the raw catalog data.

    The function loads data from a FITS file, processes it by sorting and extracting unique sources, 
    and computes additional parameters like time steps, observation IDs, and flux data.

    Returns:
    Dict: A dictionary of processed source data, with keys being the names of the sources and values being 
          the corresponding data structured in a pre-defined format.

    Note:
    - The function handles different catalogs with specific processing needs, particularly for time-based data.
    - An error in data loading is handled and reported.
    """
    
    print(f"Loading {cat}...")
    try:
        with fits.open(file_to_load, memmap=True) as raw_data:
            sources_raw = Table(raw_data[1].data)
            sources_raw = sources_raw[np.argsort(sources_raw[dict_cat.src_names[cat]])]
    except Exception as error:
        print(f"An error occured : {colored(error, 'magenta')}")
        
    indices_for_source = [i for i in range(1, len(sources_raw)) if (sources_raw[dict_cat.src_names[cat]][i] != sources_raw[dict_cat.src_names[cat]][i - 1])]

    if cat == "Swift":
        time_start_obs = Time(sources_raw["StartTime_UTC"], format="iso").mjd
        time_end_obs = Time(sources_raw["StopTime_UTC"], format="iso").mjd
        time_start_obs = np.split(time_start_obs, indices_for_source)
        time_end_obs = np.split(time_end_obs, indices_for_source)
        
    time_steps = np.split(np.array(sources_raw[dict_cat.dictionary_catalog[cat]["time_name"]]), indices_for_source)
    if cat in ("XMM", "Swift", "Stacked"):
        obs_ids = np.split(np.array(sources_raw[dict_cat.dictionary_catalog[cat]["obsid_name"]]), indices_for_source)
    else:
        obs_ids = [[] for elt in indices_for_source]
    names = np.split(np.array(sources_raw[dict_cat.src_names[cat]]), indices_for_source)

    band_flux, band_flux_errors_pos, band_flux_errors_neg = [], [], []
    
    flux = np.split(dict_cat.dictionary_catalog[cat]["conv_factor"]*np.array(sources_raw[dict_cat.dictionary_catalog[cat]["flux_obs"]]), indices_for_source)
    flux_errors_neg = np.split(dict_cat.dictionary_catalog[cat]["conv_factor"]*np.array(sources_raw[dict_cat.dictionary_catalog[cat]["flux_obs_err"][0]]), indices_for_source)
    flux_errors_pos = np.split(dict_cat.dictionary_catalog[cat]["conv_factor"]*np.array(sources_raw[dict_cat.dictionary_catalog[cat]["flux_obs_err"][1]]), indices_for_source)
    flux_errors = [[flux_neg, flux_pos] for (flux_neg, flux_pos) in zip(flux_errors_neg, flux_errors_pos)]

    band_flux_obs = dict_cat.dictionary_catalog[cat]["band_flux_obs"] #band_flux_obs_err
    band_flux_obs_err_neg, band_flux_obs_err_pos = dict_cat.dictionary_catalog[cat]["band_flux_obs_err"][0], dict_cat.dictionary_catalog[cat]["band_flux_obs_err"][1]
    for band_flux_name, band_flux_err_neg_name, band_flux_err_pos_name in zip(band_flux_obs, band_flux_obs_err_neg, band_flux_obs_err_pos):
        band_flux.append(np.array(sources_raw[band_flux_name]))
        band_flux_errors_pos.append(np.array(sources_raw[band_flux_err_pos_name]))
        band_flux_errors_neg.append(np.array(sources_raw[band_flux_err_neg_name]))
        
    band_flux = np.transpose(np.array(band_flux))
    band_flux_errors_pos = np.transpose(np.array(band_flux_errors_pos))
    band_flux_errors_neg = np.transpose(np.array(band_flux_errors_neg))
    band_flux = np.split(band_flux, indices_for_source)
    band_flux_errors_pos = np.split(band_flux_errors_pos, indices_for_source)
    band_flux_errors_neg = np.split(band_flux_errors_neg, indices_for_source)

    band_flux_err = [[flux_neg, flux_pos] for (flux_neg, flux_pos) in zip(band_flux_errors_neg, band_flux_errors_pos)]
    dict_sources = {}
    
    #This loops on all sources, to build the Source objects
    for (index, flux, flux_error, time, name, band_flux, band_flux_err, obsid) in zip(range(len(flux)),flux, flux_errors, time_steps, names, band_flux, band_flux_err, obs_ids):
            swift_stacked_flux=[]
            swift_stacked_flux_err=[[],[]]
            swift_stacked_times=[[],[]]
            if cat == "Swift":
                tab_src_timestartobs = time_start_obs[index]
                tab_src_timeendobs = time_end_obs[index]

                #We select the stacked Swift detections first
                swift_stacked_flux=flux[obsid>1e10]
                swift_stacked_flux_err=[flux_error[0][obsid>1e10],flux_error[1][obsid>1e10]]
                swift_stacked_times=[tab_src_timestartobs[obsid>1e10], tab_src_timeendobs[obsid>1e10]]

                # We then treat the classical, non-stacked Swift detections
                flux = flux[obsid < 1e10]
                flux_error = [flux_error[0][obsid < 1e10], flux_error[1][obsid < 1e10]]
                time = time[np.where(obsid < 1e10)]
                band_flux = band_flux[obsid < 1e10]
                band_flux_err = [band_flux_err[0][obsid < 1e10], band_flux_err[1][obsid < 1e10]]
                obsid = obsid[obsid < 1e10]

            band_data = catalog_class.BandFlux(flux=band_flux, flux_err=band_flux_err)
            swift_data = catalog_class.SwiftData(stacked_flux=swift_stacked_flux, stacked_flux_err=swift_stacked_flux_err, stacked_times=swift_stacked_times)
            source = catalog_class.Source(catalog=cat, iau_name=name[0].strip(), flux=flux, flux_err=flux_error, time_steps=time, 
                            band_flux_data=band_data, obsids=[...], swift_data=swift_data, xmm_offaxis=[], short_term_var=[])
            # **kwargs : obsids, swift_data, xmm_offaxis, short_term_var
            
            dict_sources[name[0].strip()] = source
            
    return dict_sources


def load_master_sources(file_to_load: str) -> Dict:
    """
    Loads multi-instrument source data from a master source file, integrating data from various catalogs.

    Args:
    file_to_load (str): Directory path where the master source file and related catalog files are located.

    The function reads a master source cone FITS file and integrates it with relevant catalog data. 
    It combines data from multiple catalogs to create a comprehensive view of each master source.

    Returns:
    Dict: A dictionary of master sources, where each key is a master source ID and its value is an object 
          representing the combined data from various catalogs.

    Note:
    - The function assumes the existence of catalog-specific FITS files in the same directory as the master source file.
    - Any errors during data loading are handled and reported.
    """
    
    """Loads the multi-instruments sources in a dictionary"""
    print(f"Loading Master Sources...")
    path_file_to_load = os.path.join(file_to_load, 'Master_source_cone.fits').replace("\\", "/")
    try:
        with fits.open(path_file_to_load, memmap=True) as raw_data:
            sources_raw = Table(raw_data[1].data)
    except Exception as error:
        print(f"An error occured : {colored(error, 'magenta')}")
        
    tab_catalog_sources = {}
    for cat in dict_cat.catalogs:
        catalog_path = os.path.join(file_to_load, cat+'.fits').replace("\\", "/")
        try:
            tab_catalog_sources[cat] = load_relevant_sources(cat, catalog_path)
        except Exception as error:
            print(f"No sources detected in {cat} catalog !")
        
    dict_master_sources = {}
    for line in tqdm(sources_raw):
        tab_sources_for_this_ms = []
        for cat in dict_cat.catalogs:
            try:
                if line[cat] != '':
                    name = line[cat].strip()
                    if name in tab_catalog_sources[cat].keys():
                        tab_sources_for_this_ms.append(tab_catalog_sources[cat][name])
            except Exception as error:
                pass
        ms_id = line["MS_ID"]
        ms = catalog_class.MasterSource(ms_id, tab_sources_for_this_ms, line["MS_RA"], line["MS_DEC"], line["MS_POSERR"])
        dict_master_sources[ms_id] = ms
        
    print("Master sources loaded!")
    return dict_master_sources


def master_source_plot(master_sources: Dict, simulation_data: Dict, number_graph: int) -> None:
    """
    Generates and saves plots for multi-instrument sources based on catalog data.

    Args:
    master_sources (Dict): A dictionary of multi-instrument sources to be plotted.
    simulation_data (Dict): A dictionary containing simulation data including object positions.
    number_graph (int): The number of graphs to generate from the master sources.

    The function iterates over a specified number of master sources and generates a plot for each. 
    It combines data from different catalogs and plots energy bands and fluxes.

    Note:
    - The function is designed to handle various catalogs and adjusts the plot according to the specific data available for each source.
    - The plots are saved in a specified directory, and their names are indexed.
    """

    object_data = simulation_data["object_data"]
    plot_var_sources_path = simulation_data["os_dictionary"]["plot_var_sources_path"]
    
    count = 0
    
    for multi_instrument_source in list(master_sources.values())[:number_graph]:
        #Each multi_instrument_source is an object with the underlying catalog sources associated with it

        #Here we compute the NICER off-axis angle between the source and the pointing
        source_coords = SkyCoord(multi_instrument_source.ra*u.degree, multi_instrument_source.dec*u.degree, frame="icrs")
        off_axis = object_data["object_position"].separation(source_coords)

        # plt.figure(figsize=(15, 8))
        figure, axes = plt.subplots(1, 1, figsize=(15, 8))
        for catalog in multi_instrument_source.sources.keys():
            #If a given catalog is contained in this source, it will be in the "sources" dictionary, catalog as key,
            #source object as value
            catalog_source = multi_instrument_source.sources[catalog]
            tab_width = 2 * np.array(dict_cat.dictionary_catalog[catalog]["energy_band_half_width"])
            for band_det in range(len(catalog_source.band_flux)):
                #The band fluxes are stored in catalog_source.band_flux. They're in erg/s/cm2, so divide by tab_width to
                #be in erg/s/cm2/keV. Here I plot them, but you can do whatever you want with those
                axes.step(dict_cat.band_edges[catalog], 
                        [catalog_source.band_flux[band_det][0] / tab_width[0]] 
                        + list(catalog_source.band_flux[band_det] / tab_width),
                        c=dict_cat.colors[catalog], where='pre')
                axes.errorbar(dict_cat.dictionary_catalog[catalog]["energy_band_center"], catalog_source.band_flux[band_det] / tab_width,
                              yerr=[catalog_source.band_flux_err[0][band_det] / tab_width,
                                    catalog_source.band_flux_err[1][band_det] / tab_width],
                              fmt="o", markeredgecolor='gray', c=dict_cat.colors[catalog], alpha=0.4)
            axes.step([], [], c=dict_cat.colors[catalog], label=f"{catalog_source.iau_name}, {catalog}")
        axes.set_xlabel("Energy [keV]")
        axes.set_ylabel(r"$F_{\nu}$ [$\mathrm{erg.s}^{-1}.\mathrm{cm}^{-2}.\mathrm{keV}^{-1}$]")
        axes.legend()
        axes.loglog()
        
        img_path = os.path.join(plot_var_sources_path, f'sources_plot_{count}.png')
        plt.savefig(img_path)
        plt.close()
        
        count += 1


# -------------------------------------------- #

# --------------- modeling spectra with jaxspec --------------- # 


def cross_catalog_index(output_name: str, key: str, iauname: str, nearby_sources_table: Table) -> List:
    """
    Determines the indices in a nearby sources table that correspond to sources found in a master source cone file.

    Args:
    output_name (str): Directory path where the master source cone file is located.
    key (str): Key representing the catalog (e.g., "CS_Chandra" or "Chandra").
    iauname (str): Column name in the nearby sources table representing the IAU name of sources.
    nearby_sources_table (Table): A table containing data about nearby sources.

    The function reads the master source cone FITS file to extract source names. It then matches these 
    names with those in the nearby sources table to find corresponding indices.

    Returns:
    List: A list of indices from the nearby sources table that match with the master source cone.

    Note:
    - This function is specific to astronomical data analysis where cross-catalog matching is required.
    """
    
    master_source_cone_path = os.path.join(output_name, "Master_source_cone.fits").replace("\\", "/")
    with fits.open(master_source_cone_path) as data:
        master_source_cone = Table(data[1].data)
    
    if key == "CS_Chandra":
        key = "Chandra"
    
    msc_name = [name for name in master_source_cone[key] if name != ""]
    var_index_in_nearby_sources_table = []
    for name in msc_name:
        if name in nearby_sources_table[iauname]:
            index_in_table = list(nearby_sources_table[iauname]).index(name)
            var_index_in_nearby_sources_table.append(index_in_table)
            
    return var_index_in_nearby_sources_table


def modeling_source_spectra(nearby_sources_table: Table, instrument, model, var_index) -> List:
    """
    Generates model spectra for sources in a nearby sources table using specified instrument and model parameters.

    Args:
    nearby_sources_table (Table): A table containing data of nearby sources.
    instrument: Instrument object with parameters for generating spectra.
    model: Spectral model to be applied for the spectra generation.
    var_index (List): List of indices indicating variable sources in the nearby sources table.

    The function iterates through the nearby sources table, applying the spectral model to each source. 
    It accounts for vignetting factors and adjusts parameters like 'N_H' and 'alpha' based on source data.

    Returns:
    List: A tuple of two lists - the total spectra and total variable spectra for all sources.

    Note:
    - The function assumes 'vignetting_factor', 'Nh', and 'Photon Index' columns in the nearby sources table.
    - 'fakeit_for_multiple_parameters' function is used for generating fake spectra based on the given model and instrument.
    """
    
    print(f"\n{colored('Modeling spectra...', 'yellow', attrs=['underline'])}")
    total_spectra = []
    total_var_spectra = []
    size = 10_000
    
    for index, vignet_factor in tqdm(enumerate(nearby_sources_table["vignetting_factor"])):
        parameters = {}
        parameters = {
            "tbabs_1": {"N_H": np.full(size, nearby_sources_table["Nh"][index]/1e22)},
            "powerlaw_1": {
                "alpha": np.full(size, nearby_sources_table["Photon Index"][index] if nearby_sources_table["Photon Index"][index] > 0.0 else 1.7),
                "norm": np.full(size, 1e-5),
            }
        }
        
        spectra = fakeit_for_multiple_parameters(instrument=instrument, model=model, parameters=parameters) * vignet_factor

        if index in var_index:
            total_var_spectra.append(spectra)
        
        total_spectra.append(spectra)
        
    return total_spectra, total_var_spectra


def total_plot_spectra(total_spectra: List, total_var_spectra: List, instrument, simulation_data: Dict, catalog_name: str) -> Dict:
    """
    Generates and saves a plot of spectral modeling data from a specified catalog and returns spectral data.

    Args:
    total_spectra (List): A list containing spectra from nearby sources.
    total_var_spectra (List): A list containing variability spectra data.
    instrument: An object containing instrument-specific data such as output energies.
    simulation_data (Dict): A dictionary containing simulation parameters and paths.
    catalog_name (str): Name of the catalog used for spectral modeling.

    The function creates three subplots:
    1. Spectra from Nearby Sources: Plots the median spectra for each source.
    2. Sum of Spectra: Shows the sum of all spectra.
    3. Spectrum Summed with Variability Sources Error: Includes error bars representing variability.

    The plot includes logarithmic scaling on both axes, and the x-axis represents energy in keV. 
    The y-axis represents counts. The plot is saved as a PNG file in a directory specified in `simulation_data`.

    Returns:
    Dict: A dictionary containing energy, counts, and their upper and lower limits.

    Note:
    - The saved plot is named based on the catalog name and the object of interest from the simulation data.
    - The function assumes `instrument` has an attribute `out_energies` for energy values.
    """
    
    object_data = simulation_data["object_data"]
    os_dictionary = simulation_data["os_dictionary"]
    graph_data = {"min_lim_x": 0.2,
                  "max_lim_x": 10.0,
                  "percentile_0": 10,
                  "percentile_2": 90}

    figure_1, axes = plt.subplots(1, 3, figsize=(17, 9), sharey=True)
    figure_1.suptitle(f"Spectral modeling close to {object_data['object_name']}\ncatalog : {catalog_name}", fontsize=20)
    figure_1.text(0.5, 0.04, 'Energy [keV]', ha='center', va='center', fontsize=16)
    figure_1.text(0.085, 0.5, 'Counts', ha='center', va='center', rotation='vertical', fontsize=16)

    for ax in axes:
        ax.set_xlim([graph_data["min_lim_x"], graph_data["max_lim_x"]])
        ax.loglog()

    ax0 = axes[0]
    for spectra in total_spectra:
        ax0.step(instrument.out_energies[0],
                np.median(spectra, axis=0),
                where="post")
    ax0.set_title("Spectra from Nearby Sources")

    spectrum_summed = 0.0
    for item in range(len(total_spectra)):
        spectrum_summed += total_spectra[item]

    spectrum_var_summed = 0.0
    for item in range(len(total_var_spectra)):
        spectrum_var_summed += total_var_spectra[item]  

    y_upper = np.median(spectrum_summed, axis=0) + np.median(spectrum_var_summed, axis=0)
    y_lower = np.median(spectrum_summed, axis=0) - np.median(spectrum_var_summed, axis=0)

    ax1 = axes[1]
    ax1.step(instrument.out_energies[0],
            np.median(spectrum_summed, axis=0),
            where='post', color='black'
            )

    ax1.set_title("Sum of spectra")

    ax2 = axes[2]
    ax2.errorbar(instrument.out_energies[0], y=np.median(spectrum_summed, axis=0), yerr=np.median(spectrum_var_summed, axis=0), 
                fmt="none", ecolor='red', capsize=2, capthick=3,
                label='error')
    ax2.step(instrument.out_energies[0], np.median(spectrum_summed, axis=0), color='black', label="sum powerlaw")
    ax2.set_title("Spectrum Summed with var sources error")
    ax2.legend(loc='upper right')
    ax2.loglog()

    key = simulation_data["os_dictionary"]["catalog_key"]
    img_path = os.path.join(os_dictionary['img'], f"{key}_spectral_modeling_close_to_{object_data['object_name']}.png".replace(" ", "_")).replace("\\", "/")
    plt.savefig(img_path)
    plt.show()
    
    data = {
        "Energy": instrument.out_energies[0],
        "Counts": np.median(spectrum_summed, axis=0),
        "Upper limit": y_upper,
        "Lower limit": y_lower
    }
    
    return data


def write_txt_file(simulation_data: Dict, data: Dict) -> None:
    """
    Writes the spectral modeling data into a formatted text file.

    Args:
    simulation_data (Dict): A dictionary containing simulation parameters and paths.
    data (Dict): A dictionary with spectral data including energy, counts, and their upper and lower limits.

    The function creates a text file in the specified directory within the simulation data dictionary. 
    The file includes a header and rows of data, each containing energy values, count rates, and their upper 
    and lower limits. The data is formatted for readability and analysis purposes.

    Note:
    - The output text file is named based on the 'catalog_key' value in the simulation data dictionary.
    - The function expects 'data' to have keys corresponding to 'Energy', 'Counts', 'Upper limit', 
      and 'Lower limit' and their associated values.
    """
    
    catalog_directory = simulation_data['os_dictionary']["catalog_directory"]
    key = simulation_data["os_dictionary"]["catalog_key"]
    txt_path = os.path.join(catalog_directory, f'{key}_output_modeling_plot.txt').replace("\\", "/")
    
    data_to_txt = [
        list(data.keys())
    ]
    
    energy, counts, y_upper, y_lower = list(data.values())
    data_to_txt.extend([energy[index], counts[index], y_upper[index], y_lower[index]] for index in range(len(energy)))
    
    with open(txt_path, 'w') as file:
        header = "{:<15} {:<15} {:<15} {:<15}".format(*data_to_txt[0])
        file.write(header + "\n")

        for row in data_to_txt[1:]:
            new_row = "{:<10.5f}     {:<10.5f}       {:<10.5f}       {:<10.5f}".format(*row)
            file.write(new_row + "\n")
            
    print(f"\n{colored(f'{key}_output_modeling_plot.txt', 'yellow')} has been created in {colored(txt_path, 'blue')}")


# ------------------------------------------------------------- #